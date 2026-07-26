import json
import os
import re
import sys
import tempfile
from datetime import date, datetime

from openpyxl import load_workbook

try:
    import msoffcrypto
except ImportError:
    msoffcrypto = None


def normalize_cell(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()[:10]
    return str(value).strip()


def load_plain_workbook(path):
    return load_workbook(path, read_only=True, data_only=True)


def load_decrypted_workbook(input_path, password):
    if msoffcrypto is None:
        raise RuntimeError("msoffcrypto is not installed")
    temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    temp.close()
    try:
        with open(input_path, "rb") as source:
            office_file = msoffcrypto.OfficeFile(source)
            office_file.load_key(password=password)
            with open(temp.name, "wb") as target:
                office_file.decrypt(target)
        return load_plain_workbook(temp.name), temp.name
    except Exception:
        try:
            os.unlink(temp.name)
        except OSError:
            pass
        raise


def workbook_for(path, password):
    try:
        return load_plain_workbook(path), None
    except Exception as plain_error:
        try:
            return load_decrypted_workbook(path, password)
        except Exception as decrypt_error:
            raise RuntimeError(f"plain open failed: {plain_error} / decrypt failed: {decrypt_error}") from decrypt_error


HEADER_KEYWORDS = {
    "입고요청일", "배송일", "배송일자", "확정호차", "기준호차", "호차", "톤수",
    "기사명", "연락처", "배송권역", "고객", "고객코드", "고객명", "매출금액",
    "매출액", "주문금액", "총주문금액", "총주문액", "금액", "배송건수", "중량",
    "고객주소", "상세주소", "주소", "배송주소", "순번", "순서",
}


def header_score(values):
    joined = "|".join(values)
    return sum(1 for keyword in HEADER_KEYWORDS if keyword in joined)


def infer_date_from_values(values):
    joined = " ".join(values)
    match = re.search(r"(\d{2,4})[.\-/년\s]+(\d{1,2})[.\-/월\s]+(\d{1,2})", joined)
    if not match:
        return ""
    year, month, day = match.groups()
    if len(year) == 2:
        year = f"20{year}"
    return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"


def rows_from_sheet(sheet, source_file):
    rows = []
    columns = set()
    normalized_rows = [[normalize_cell(value) for value in row_values] for row_values in sheet.iter_rows(values_only=True)]
    inferred_date = ""
    header_index = None
    best_score = 0

    for index, normalized in enumerate(normalized_rows[:120]):
        if not inferred_date:
            inferred_date = infer_date_from_values(normalized)
        score = header_score(normalized)
        if score > best_score:
            best_score = score
            header_index = index

    if header_index is None or best_score < 3:
        for index, normalized in enumerate(normalized_rows):
            if any(normalized):
                header_index = index
                break
    if header_index is None:
        return rows, columns

    header = [value or f"column_{index + 1}" for index, value in enumerate(normalized_rows[header_index])]
    date_column_exists = any(("입고요청일" in column or "배송일" in column or "배송일자" in column or "일자" in column or "날짜" in column) for column in header)
    if inferred_date and not date_column_exists:
        header.append("입고요청일")
    for column in header:
        if column and not column.startswith("__EMPTY"):
            columns.add(column)

    for normalized in normalized_rows[header_index + 1:]:
        row = {}
        for index, column in enumerate(header):
            if not column or column.startswith("__EMPTY"):
                continue
            if column == "입고요청일" and index >= len(normalized):
                row[column] = inferred_date
            else:
                row[column] = normalized[index] if index < len(normalized) else ""
        if any(row.values()):
            row["__rawValues"] = normalized[:len(header)]
            row["__headers"] = header
            row["_sourceFile"] = source_file
            row["_sourceSheet"] = sheet.title
            rows.append(row)

    return rows, columns


def main():
    if len(sys.argv) != 5:
        print("usage: parse_excel.py input output password source_name", file=sys.stderr)
        return 2

    input_path, output_path, password, source_name = sys.argv[1:5]
    decrypted_path = None
    try:
        workbook, decrypted_path = workbook_for(input_path, password)
        all_rows = []
        all_columns = set()
        for sheet in workbook.worksheets:
            rows, columns = rows_from_sheet(sheet, source_name)
            all_rows.extend(rows)
            all_columns.update(columns)
        workbook.close()

        with open(output_path, "w", encoding="utf-8") as target:
            json.dump({"rows": all_rows, "columns": sorted(all_columns)}, target, ensure_ascii=False, separators=(",", ":"))
        return 0
    finally:
        if decrypted_path:
            try:
                os.unlink(decrypted_path)
            except OSError:
                pass


if __name__ == "__main__":
    raise SystemExit(main())
